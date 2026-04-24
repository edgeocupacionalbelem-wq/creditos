
import os
import re
import zipfile
import requests
from collections import defaultdict
from datetime import datetime
from io import BytesIO

from flask import Flask, render_template, request, redirect, url_for, flash, Response, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, or_, text

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = os.environ.get("SECRET_KEY", "troque-esta-chave")
app.config["MAX_CONTENT_LENGTH"] = 80 * 1024 * 1024

db_url = os.environ.get("DATABASE_URL", "sqlite:///creditos.db")
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
ALLOWED_ATTACHMENT_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png"}

class BaseMetadata(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    original_filename = db.Column(db.String(255), nullable=False)
    updated_at = db.Column(db.String(30), nullable=False)
    months_count = db.Column(db.Integer, nullable=False, default=0)
    total_creditos = db.Column(db.Integer, nullable=False, default=0)

class Credito(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    mes_index = db.Column(db.Integer, nullable=False)
    mes_nome = db.Column(db.String(120), nullable=False, index=True)
    recibo = db.Column(db.String(120), nullable=True, index=True)
    empresa = db.Column(db.String(255), nullable=True, index=True)
    documento = db.Column(db.String(30), nullable=True)
    pix_id = db.Column(db.String(255), nullable=True, index=True)
    pagamento_info = db.Column(db.Text, nullable=True)
    credito_count = db.Column(db.Integer, nullable=False, default=1)

class Comprovante(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    recibo = db.Column(db.String(120), nullable=False, index=True)
    pix_id = db.Column(db.String(255), nullable=False, index=True)
    slot = db.Column(db.Integer, nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    cloudinary_url = db.Column(db.Text, nullable=False)
    cloudinary_public_id = db.Column(db.String(255), nullable=False)
    resource_type = db.Column(db.String(50), nullable=False, default="auto")

with app.app_context():
    db.create_all()
    try:
        db.session.execute(text("ALTER TABLE credito ADD COLUMN pix_id VARCHAR(255)"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE credito ADD COLUMN pagamento_info TEXT"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_credito_recibo_pix ON credito (recibo, pix_id)"))
        db.session.commit()
    except Exception:
        db.session.rollback()

try:
    import cloudinary
    import cloudinary.uploader
    cloudinary.config(
        cloud_name=os.environ.get("CLOUDINARY_CLOUD_NAME"),
        api_key=os.environ.get("CLOUDINARY_API_KEY"),
        api_secret=os.environ.get("CLOUDINARY_API_SECRET"),
        secure=True,
    )
except Exception:
    cloudinary = None

@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response

@app.route("/favicon.ico")
def favicon():
    return Response(status=204)

def cloudinary_ready():
    return all([
        os.environ.get("CLOUDINARY_CLOUD_NAME"),
        os.environ.get("CLOUDINARY_API_KEY"),
        os.environ.get("CLOUDINARY_API_SECRET"),
        cloudinary is not None
    ])

def allowed_file(filename: str) -> bool:
    return os.path.splitext(filename.lower())[1] in ALLOWED_EXTENSIONS

def attachment_allowed(filename):
    return os.path.splitext(filename.lower())[1] in ALLOWED_ATTACHMENT_EXTENSIONS

def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()

def normalize_rows(rows):
    return [[clean_cell(v) for v in row] for row in rows]

def normalize_header_name(text: str) -> str:
    t = clean_cell(text).upper()
    t = t.replace("Nº", "NO").replace("N°", "NO")
    t = re.sub(r"[^A-Z0-9 ]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

def detect_header_and_rows(rows):
    rows = normalize_rows(rows)
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:15]):
        normalized = [normalize_header_name(c) for c in row]
        keywords = ["RECIBO", "SETOR", "STATUS", "EXAME", "FUNCIONARIO", "DATA", "TIPO"]
        keyword_hits = sum(1 for c in normalized if any(k in c for k in keywords))
        non_empty = sum(1 for c in row if c)
        score = keyword_hits * 8 + non_empty
        if score > best_score:
            best_score = score
            best_idx = idx
    header = rows[best_idx] if rows else []
    data = rows[best_idx + 1:] if best_idx + 1 < len(rows) else []
    while data and not any(data[-1]):
        data.pop()
    return header, data

def find_column_index(header, aliases, fallback_index=None):
    normalized = [normalize_header_name(c) for c in header]
    for alias in aliases:
        alias_norm = normalize_header_name(alias)
        for idx, col in enumerate(normalized):
            if alias_norm and alias_norm in col:
                return idx
    if fallback_index is not None and fallback_index < len(header):
        return fallback_index
    return None

def format_document(doc):
    digits = re.sub(r"\D", "", doc or "")
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return clean_cell(doc)

def parse_setor(setor_text):
    raw = clean_cell(setor_text)
    if not raw:
        return "", ""
    doc_match = re.search(r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}|\d{3}\.?\d{3}\.?\d{3}-?\d{2})', raw)
    doc = doc_match.group(1) if doc_match else ""
    doc = format_document(doc)
    empresa = raw
    if doc_match:
        empresa = (raw[:doc_match.start()] + raw[doc_match.end():]).strip(" -–•|")
    empresa = re.sub(r"\s+", " ", empresa).strip(" -–•|")
    return empresa, doc

def detect_credit_status(row, status_idx_primary, status_idx_secondary):
    values = []
    if status_idx_primary is not None and status_idx_primary < len(row):
        values.append(clean_cell(row[status_idx_primary]).upper())
    if status_idx_secondary is not None and status_idx_secondary < len(row):
        values.append(clean_cell(row[status_idx_secondary]).upper())
    return any(v == "NÃO REALIZADO" or v == "NAO REALIZADO" for v in values)

def get_pix_col_index(mes_nome):
    nome = clean_cell(mes_nome).upper()
    if "JANEIRO" in nome:
        return 7
    return 1

def process_workbook(filepath):
    # read_only drastically reduces memory and style processing
    wb = load_workbook(filepath, data_only=True, read_only=True, keep_links=False)
    months = []
    total_creditos = 0
    for ws_idx, ws in enumerate(wb.worksheets, start=1):
        raw_rows = list(ws.iter_rows(values_only=True))
        header, data = detect_header_and_rows(raw_rows)
        if not header:
            continue
        recibo_idx = find_column_index(header, ["RECIBO", "NO RECIBO", "NUMERO RECIBO"], fallback_index=2)
        setor_idx = find_column_index(header, ["SETOR"], fallback_index=11)
        status_h_idx = 7 if len(header) > 7 else None
        status_d_idx = 3 if len(header) > 3 else None
        pix_idx = get_pix_col_index(ws.title)
        pagamento_info_idx = 8 if len(header) > 8 else None

        grouped = defaultdict(lambda: {"credito_count": 0, "recibo": "", "empresa": "", "documento": "", "pix_id": "", "pagamento_info": ""})
        for row in normalize_rows(data):
            if not any(row):
                continue
            if not detect_credit_status(row, status_h_idx, status_d_idx):
                continue
            recibo = clean_cell(row[recibo_idx]) if recibo_idx is not None and recibo_idx < len(row) else ""
            setor = clean_cell(row[setor_idx]) if setor_idx is not None and setor_idx < len(row) else ""
            pix_id = clean_cell(row[pix_idx]) if pix_idx is not None and pix_idx < len(row) else ""
            pagamento_info = clean_cell(row[pagamento_info_idx]) if pagamento_info_idx is not None and pagamento_info_idx < len(row) else ""
            empresa, documento = parse_setor(setor)
            key = (recibo, pix_id)
            grouped[key]["credito_count"] += 1
            grouped[key]["recibo"] = recibo
            grouped[key]["empresa"] = empresa
            grouped[key]["documento"] = documento
            grouped[key]["pix_id"] = pix_id
            grouped[key]["pagamento_info"] = pagamento_info
        creditos = list(grouped.values())
        creditos.sort(key=lambda x: (x["empresa"], x["recibo"]))
        mes_total = sum(item["credito_count"] for item in creditos)
        total_creditos += mes_total
        months.append({"mes_index": ws_idx, "mes_nome": ws.title, "credito_receipts_count": len(creditos), "credito_total_count": mes_total, "creditos": creditos})
    wb.close()
    return months, total_creditos

def credito_key(recibo, pix_id):
    return (clean_cell(recibo), clean_cell(pix_id))

def cleanup_orphan_attachments(valid_keys):
    for comp in Comprovante.query.all():
        if credito_key(comp.recibo, comp.pix_id) not in valid_keys:
            if cloudinary_ready():
                try:
                    cloudinary.uploader.destroy(comp.cloudinary_public_id, resource_type=comp.resource_type, invalidate=True)
                except Exception:
                    pass
            db.session.delete(comp)
    db.session.commit()

def replace_database_with_planilha(filepath, original_filename):
    months, total_creditos = process_workbook(filepath)
    valid_keys = set()
    for month in months:
        for item in month["creditos"]:
            valid_keys.add(credito_key(item["recibo"], item["pix_id"]))
    db.session.query(Credito).delete()
    db.session.query(BaseMetadata).delete()
    db.session.add(BaseMetadata(
        original_filename=original_filename,
        updated_at=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        months_count=len(months),
        total_creditos=total_creditos,
    ))
    for month in months:
        for item in month["creditos"]:
            db.session.add(Credito(
                mes_index=month["mes_index"],
                mes_nome=month["mes_nome"],
                recibo=item["recibo"],
                empresa=item["empresa"],
                documento=item["documento"],
                pix_id=item["pix_id"],
                pagamento_info=item["pagamento_info"],
                credito_count=item["credito_count"],
            ))
    db.session.commit()
    cleanup_orphan_attachments(valid_keys)


def get_dashboard_data(start_idx=None, end_idx=None, busca="", mes_nome="", empresa_filtro="", status_filtro="todos"):
    meta = BaseMetadata.query.order_by(BaseMetadata.id.desc()).first()

    month_rows = db.session.query(
        Credito.mes_index,
        Credito.mes_nome,
        func.count(Credito.id).label("credito_receipts_count"),
        func.coalesce(func.sum(Credito.credito_count), 0).label("credito_total_count"),
    ).group_by(Credito.mes_index, Credito.mes_nome).order_by(Credito.mes_index.asc()).all()

    months_summary = [
        {
            "mes_index": r.mes_index,
            "mes_nome": r.mes_nome,
            "credito_receipts_count": int(r.credito_receipts_count or 0),
            "credito_total_count": int(r.credito_total_count or 0),
        }
        for r in month_rows
    ]

    if months_summary:
        min_idx = months_summary[0]["mes_index"]
        max_idx = months_summary[-1]["mes_index"]
    else:
        min_idx = max_idx = 1

    mes_nome = (mes_nome or "").strip()
    if mes_nome:
        selected = next((m for m in months_summary if m["mes_nome"] == mes_nome), None)
        if selected:
            start_idx = selected["mes_index"]
            end_idx = selected["mes_index"]

    if start_idx is None:
        start_idx = min_idx
    if end_idx is None:
        end_idx = max_idx
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx

    q = Credito.query.filter(Credito.mes_index >= start_idx, Credito.mes_index <= end_idx)

    busca = (busca or "").strip()
    if busca:
        term = f"%{busca}%"
        q = q.filter(or_(
            Credito.mes_nome.ilike(term),
            Credito.recibo.ilike(term),
            Credito.empresa.ilike(term),
            Credito.documento.ilike(term),
            Credito.pix_id.ilike(term),
            Credito.pagamento_info.ilike(term),
        ))

    empresa_filtro = (empresa_filtro or "").strip()
    if empresa_filtro:
        q = q.filter(Credito.empresa.ilike(f"%{empresa_filtro}%"))

    creditos = q.order_by(Credito.mes_index.asc(), Credito.empresa.asc(), Credito.recibo.asc()).all()

    comp_map = {}
    for c in Comprovante.query.all():
        comp_map.setdefault(credito_key(c.recibo, c.pix_id), {})[c.slot] = c

    months_map = {m["mes_index"]: {**m, "creditos": []} for m in months_summary}
    total_com_comprovante = 0
    total_parcial = 0
    total_sem_comprovante = 0

    for c in creditos:
        c.comprovantes = comp_map.get(credito_key(c.recibo, c.pix_id), {})
        comp_count = len(c.comprovantes)
        c.comprovante_count = comp_count

        if comp_count >= 2:
            c.status_comprovante = "completo"
            c.status_label = "Completo"
            total_com_comprovante += 1
        elif comp_count == 1:
            c.status_comprovante = "parcial"
            c.status_label = "Parcial"
            total_parcial += 1
        else:
            c.status_comprovante = "sem"
            c.status_label = "Sem comprovante"
            total_sem_comprovante += 1

        if status_filtro == "completo" and comp_count < 2:
            continue
        if status_filtro == "parcial" and comp_count != 1:
            continue
        if status_filtro == "sem" and comp_count != 0:
            continue
        if status_filtro == "com_algum" and comp_count == 0:
            continue

        if c.mes_index in months_map:
            months_map[c.mes_index]["creditos"].append(c)

    months_filtered = [months_map[idx] for idx in sorted(months_map) if start_idx <= idx <= end_idx]
    resumo_total = sum(sum(c.credito_count for c in m["creditos"]) for m in months_filtered)
    resumo_recibos = sum(len(m["creditos"]) for m in months_filtered)

    empresas = [
        row[0]
        for row in db.session.query(Credito.empresa)
        .filter(Credito.empresa.isnot(None), Credito.empresa != "")
        .distinct()
        .order_by(Credito.empresa.asc())
        .all()
    ]

    logs = []
    if meta:
        logs.append({"titulo": "Base ativa", "descricao": meta.original_filename, "data": meta.updated_at})
    ultimo_comprovante = Comprovante.query.order_by(Comprovante.id.desc()).first()
    if ultimo_comprovante:
        logs.append({
            "titulo": "Último comprovante anexado",
            "descricao": f"Recibo {ultimo_comprovante.recibo} • Slot {ultimo_comprovante.slot}",
            "data": ultimo_comprovante.original_filename,
        })

    return {
        "meta": meta,
        "months_summary": months_summary,
        "months_filtered": months_filtered,
        "resumo_total": resumo_total,
        "resumo_recibos": resumo_recibos,
        "min_idx": min_idx,
        "max_idx": max_idx,
        "inicio": start_idx,
        "fim": end_idx,
        "busca": busca,
        "mes_nome": mes_nome,
        "empresa_filtro": empresa_filtro,
        "status_filtro": status_filtro,
        "empresas": empresas,
        "logs": logs,
        "total_com_comprovante": total_com_comprovante,
        "total_parcial": total_parcial,
        "total_sem_comprovante": total_sem_comprovante,
    }


@app.route("/")
def index():
    try:
        inicio = int(request.args.get("inicio", "0")) or None
    except ValueError:
        inicio = None
    try:
        fim = int(request.args.get("fim", "0")) or None
    except ValueError:
        fim = None

    busca = request.args.get("busca", "")
    mes_nome = request.args.get("mes", "")
    empresa_filtro = request.args.get("empresa", "")
    status_filtro = request.args.get("status", "todos")

    css_path = os.path.join(app.static_folder, "styles.css")
    with open(css_path, encoding="utf-8") as f:
        inline_css = f.read()

    data = get_dashboard_data(
        start_idx=inicio,
        end_idx=fim,
        busca=busca,
        mes_nome=mes_nome,
        empresa_filtro=empresa_filtro,
        status_filtro=status_filtro,
    )
    data["inline_css"] = inline_css
    return render_template("index.html", **data)

@app.route("/upload-base", methods=["POST"])
def upload_base():
    file = request.files.get("base_file")
    if not file or not file.filename:
        flash("Selecione uma planilha .xlsx ou .xlsm.")
        return redirect(url_for("index"))
    if not allowed_file(file.filename):
        flash("Formato inválido. Envie apenas .xlsx ou .xlsm.")
        return redirect(url_for("index"))
    filename = secure_filename(file.filename)
    temp_dir = os.environ.get("TMPDIR", "/tmp")
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, filename)
    file.save(temp_path)
    try:
        if os.path.getsize(temp_path) > 10_000_000:
            flash("Planilha muito pesada. Tente reduzir o tamanho ou remover formatações.")
            return redirect(url_for("index"))
        replace_database_with_planilha(temp_path, file.filename)
        flash("Base atualizada com sucesso.")
    except Exception as e:
        flash(f"Não foi possível ler a planilha: {e}")
    finally:
        try: os.remove(temp_path)
        except OSError: pass
    return redirect(url_for("index"))

@app.route("/upload-comprovante/<path:recibo>/<path:pix_id>/<int:slot>", methods=["POST"])
def upload_comprovante(recibo, pix_id, slot):
    inicio = request.form.get("inicio", "")
    fim = request.form.get("fim", "")
    busca = request.form.get("busca", "")
    mes = request.form.get("mes", "")
    empresa = request.form.get("empresa", "")
    status = request.form.get("status", "todos")
    file = request.files.get("arquivo")
    if slot not in (1, 2):
        flash("Slot inválido.")
        return redirect(url_for("index", inicio=inicio, fim=fim, busca=busca, mes=mes, empresa=empresa, status=status))
    if not file or not file.filename:
        flash("Selecione um comprovante.")
        return redirect(url_for("index", inicio=inicio, fim=fim, busca=busca, mes=mes, empresa=empresa, status=status))
    if not attachment_allowed(file.filename):
        flash("Envie PDF, JPG, JPEG ou PNG.")
        return redirect(url_for("index", inicio=inicio, fim=fim, busca=busca, mes=mes, empresa=empresa, status=status))
    if not cloudinary_ready():
        flash("Cloudinary não configurado no Render.")
        return redirect(url_for("index", inicio=inicio, fim=fim, busca=busca, mes=mes, empresa=empresa, status=status))
    existing = Comprovante.query.filter_by(recibo=recibo, pix_id=pix_id, slot=slot).first()
    if existing:
        try:
            cloudinary.uploader.destroy(existing.cloudinary_public_id, resource_type=existing.resource_type, invalidate=True)
        except Exception:
            pass
        db.session.delete(existing)
        db.session.commit()
    try:
        ext = os.path.splitext(file.filename.lower())[1]
        is_pdf = ext == ".pdf"
        resource_type = "raw" if is_pdf else "image"

        safe_public_base = re.sub(
            r"[^A-Za-z0-9_\\-]",
            "_",
            f"{clean_cell(recibo)}__{clean_cell(pix_id)}__slot{slot}"
        )
        public_id = f"{safe_public_base}.pdf" if is_pdf else safe_public_base

        result = cloudinary.uploader.upload(
            file,
            folder="creditos_comprovantes",
            resource_type=resource_type,
            public_id=public_id,
            overwrite=True,
            use_filename=False,
            unique_filename=False,
        )

        comp = Comprovante(
            recibo=recibo,
            pix_id=pix_id,
            slot=slot,
            original_filename=file.filename,
            cloudinary_url=result.get("secure_url") or result.get("url"),
            cloudinary_public_id=result["public_id"],
            resource_type=result.get("resource_type", resource_type),
        )
        db.session.add(comp)
        db.session.commit()
        flash("Comprovante enviado com sucesso.")
    except Exception as e:
        db.session.rollback()
        flash(f"Falha ao enviar comprovante: {e}")
    return redirect(url_for("index", inicio=inicio, fim=fim, busca=busca, mes=mes, empresa=empresa, status=status))

@app.route("/delete-comprovante/<path:recibo>/<path:pix_id>/<int:slot>", methods=["POST"])
def delete_comprovante(recibo, pix_id, slot):
    inicio = request.form.get("inicio", "")
    fim = request.form.get("fim", "")
    busca = request.form.get("busca", "")
    mes = request.form.get("mes", "")
    empresa = request.form.get("empresa", "")
    status = request.form.get("status", "todos")
    comp = Comprovante.query.filter_by(recibo=recibo, pix_id=pix_id, slot=slot).first()
    if comp:
        if cloudinary_ready():
            try:
                cloudinary.uploader.destroy(comp.cloudinary_public_id, resource_type=comp.resource_type, invalidate=True)
            except Exception:
                pass
        db.session.delete(comp)
        db.session.commit()
        flash("Comprovante apagado.")
    return redirect(url_for("index", inicio=inicio, fim=fim, busca=busca, mes=mes, empresa=empresa, status=status))


@app.route("/download-comprovante/<int:comp_id>")
def download_comprovante(comp_id):
    comp = Comprovante.query.get_or_404(comp_id)
    try:
        r = requests.get(comp.cloudinary_url, timeout=30)
        r.raise_for_status()
        data = BytesIO(r.content)
        data.seek(0)

        filename = comp.original_filename or f"comprovante_{comp.recibo}_{comp.slot}"
        mimetype = "application/pdf" if filename.lower().endswith(".pdf") else None

        return send_file(
            data,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f"Não foi possível baixar o comprovante: {e}")
        return redirect(url_for("index"))

@app.route("/abrir-comprovante/<int:comp_id>")
def abrir_comprovante(comp_id):
    comp = Comprovante.query.get_or_404(comp_id)
    return redirect(comp.cloudinary_url)

@app.route("/healthz")
def healthz():
    return {"status": "ok"}

@app.errorhandler(404)
def not_found(_):
    return redirect(url_for("index"))
